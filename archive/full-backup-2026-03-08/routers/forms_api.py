"""
Coach Forms API — Pre-built templates, AI form creator, PDF/Excel generation

8 system form templates + coach-custom forms via Azure OpenAI.
"""

import io
import json
import logging
import os
from datetime import datetime, timezone
from typing import Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.services.api_server import get_current_user, require_coach
from app.services.nate_ai_config import NATE_CHAT_URL, NATE_CHAT_KEY, nate_chat_headers, nate_chat_payload

logger = logging.getLogger("nate.forms_api")

router = APIRouter(
    prefix="/api/coach/forms",
    tags=["coach-forms"],
    dependencies=[Depends(require_coach)],
)


SYSTEM_TEMPLATES = [
    {
        "id": "privacy_policy",
        "title": "Privacy Policy",
        "description": "Data collection/storage practices, AI observation disclosure, session recording consent",
        "form_type": "system",
        "form_schema": {
            "sections": [
                {"title": "Data Collection", "fields": [
                    {"name": "data_types_collected", "label": "Types of Data Collected", "type": "info",
                     "value": "Personal information, session recordings, AI-generated insights, voice biometrics, coherence metrics"},
                    {"name": "little_nate_disclosure", "label": "AI Observation Disclosure", "type": "info",
                     "value": "Little Nate, our AI companion, observes and learns from coaching sessions to provide personalized insights."},
                ]},
                {"title": "Session Recording", "fields": [
                    {"name": "recording_consent", "label": "I consent to session recording", "type": "checkbox"},
                    {"name": "ai_observation_consent", "label": "I consent to AI observation during sessions", "type": "checkbox"},
                ]},
                {"title": "Data Retention", "fields": [
                    {"name": "video_retention", "label": "Video Retention Period", "type": "info", "value": "7 days"},
                    {"name": "transcript_retention", "label": "Transcript Retention", "type": "info", "value": "Permanent (anonymized)"},
                ]},
                {"title": "Third-Party Services", "fields": [
                    {"name": "third_parties", "label": "Services Used", "type": "info",
                     "value": "Stripe (payments), Twilio (SMS), SendGrid (email), Azure OpenAI (AI processing)"},
                ]},
                {"title": "Client Rights", "fields": [
                    {"name": "right_access", "label": "Right to Access", "type": "info", "value": "You may request a copy of all data we hold about you."},
                    {"name": "right_correction", "label": "Right to Correction", "type": "info", "value": "You may request correction of inaccurate data."},
                    {"name": "right_deletion", "label": "Right to Deletion", "type": "info", "value": "You may request deletion of your data (subject to legal retention requirements)."},
                ]},
                {"title": "Consent", "fields": [
                    {"name": "client_name", "label": "Full Legal Name", "type": "text", "required": True},
                    {"name": "consent_signature", "label": "Signature", "type": "signature", "required": True},
                    {"name": "consent_date", "label": "Date", "type": "date", "required": True},
                ]},
            ]
        },
    },
    {
        "id": "terms_of_service",
        "title": "Terms of Service",
        "description": "Platform usage rules, subscription tiers, session booking/cancellation policy",
        "form_type": "system",
        "form_schema": {
            "sections": [
                {"title": "Platform Usage", "fields": [
                    {"name": "age_requirement", "label": "Age Requirement", "type": "info", "value": "Must be 18+ (dependents allowed for minors with guardian consent)"},
                    {"name": "prohibited_conduct", "label": "Prohibited Conduct", "type": "info", "value": "Harassment, misuse of platform, sharing credentials, recording sessions without consent"},
                ]},
                {"title": "Subscription Tiers", "fields": [
                    {"name": "trial_tier", "label": "Threshold (Trial)", "type": "info", "value": "Free — limited features"},
                    {"name": "standard_tier", "label": "Inner Chamber (Standard)", "type": "info", "value": "$49/month — full access"},
                    {"name": "sovereign_tier", "label": "Sovereign Circle (Top)", "type": "info", "value": "$149/month — premium features"},
                ]},
                {"title": "Session Policy", "fields": [
                    {"name": "payment_timing", "label": "Payment Due", "type": "info", "value": "72 hours before scheduled session"},
                    {"name": "cancellation_window", "label": "Cancellation", "type": "info", "value": "24 hours before session for full refund"},
                    {"name": "no_show", "label": "No-Show Policy", "type": "info", "value": "Non-refundable after 24-hour window"},
                ]},
                {"title": "Acknowledgment", "fields": [
                    {"name": "tos_accepted", "label": "I have read and agree to these Terms of Service", "type": "checkbox", "required": True},
                    {"name": "client_name", "label": "Full Legal Name", "type": "text", "required": True},
                    {"name": "tos_date", "label": "Date", "type": "date", "required": True},
                ]},
            ]
        },
    },
    {
        "id": "adult_intake",
        "title": "Adult Intake Form",
        "description": "Demographics, emergency contact, presenting concerns, therapy history",
        "form_type": "system",
        "form_schema": {
            "sections": [
                {"title": "Personal Information", "fields": [
                    {"name": "legal_name", "label": "Full Legal Name", "type": "text", "required": True},
                    {"name": "preferred_name", "label": "Preferred Name/Pronouns", "type": "text"},
                    {"name": "dob", "label": "Date of Birth", "type": "date", "required": True},
                    {"name": "phone", "label": "Phone Number", "type": "text", "required": True},
                    {"name": "email", "label": "Email Address", "type": "text", "required": True},
                    {"name": "address", "label": "Mailing Address", "type": "textarea"},
                ]},
                {"title": "Emergency Contact", "fields": [
                    {"name": "emergency_name", "label": "Emergency Contact Name", "type": "text", "required": True},
                    {"name": "emergency_phone", "label": "Emergency Contact Phone", "type": "text", "required": True},
                    {"name": "emergency_relation", "label": "Relationship", "type": "text"},
                ]},
                {"title": "Living Situation", "fields": [
                    {"name": "living_situation", "label": "Current Living Situation", "type": "dropdown",
                     "options": ["Alone", "With partner/spouse", "With family", "With roommates", "Other"]},
                    {"name": "relationship_status", "label": "Relationship Status", "type": "dropdown",
                     "options": ["Single", "In a relationship", "Married", "Separated", "Divorced", "Widowed"]},
                    {"name": "children", "label": "Children (number and ages)", "type": "text"},
                ]},
                {"title": "Presenting Concerns", "fields": [
                    {"name": "referral_source", "label": "How did you hear about us?", "type": "text"},
                    {"name": "current_concerns", "label": "What brings you to coaching?", "type": "textarea", "required": True},
                    {"name": "previous_therapy", "label": "Previous therapy/coaching experience", "type": "textarea"},
                    {"name": "current_support", "label": "Current support system", "type": "textarea"},
                    {"name": "cultural_spiritual", "label": "Cultural or spiritual considerations", "type": "textarea"},
                ]},
                {"title": "Consent", "fields": [
                    {"name": "intake_signature", "label": "Signature", "type": "signature", "required": True},
                    {"name": "intake_date", "label": "Date", "type": "date", "required": True},
                ]},
            ]
        },
    },
    {
        "id": "medications",
        "title": "Medications Form",
        "description": "Current medications, supplements, allergies, prescriber contact",
        "form_type": "system",
        "form_schema": {
            "sections": [
                {"title": "Current Medications", "fields": [
                    {"name": "medications", "label": "List all current medications (name, dosage, frequency, prescriber)", "type": "textarea", "required": True},
                    {"name": "supplements", "label": "OTC supplements and vitamins", "type": "textarea"},
                    {"name": "recent_changes", "label": "Any medication changes in the last 90 days?", "type": "textarea"},
                ]},
                {"title": "Allergies", "fields": [
                    {"name": "allergies", "label": "Known allergies or adverse reactions", "type": "textarea"},
                ]},
                {"title": "Prescriber Information", "fields": [
                    {"name": "prescriber_name", "label": "Primary Prescribing Physician", "type": "text"},
                    {"name": "prescriber_phone", "label": "Prescriber Phone", "type": "text"},
                    {"name": "pharmacy_name", "label": "Pharmacy Name", "type": "text"},
                    {"name": "pharmacy_phone", "label": "Pharmacy Phone", "type": "text"},
                ]},
                {"title": "Substance Use", "fields": [
                    {"name": "substance_history", "label": "Substance use history (type, frequency, amount)", "type": "textarea"},
                ]},
                {"title": "Authorization", "fields": [
                    {"name": "prescriber_auth", "label": "I authorize coordination with my prescriber", "type": "checkbox"},
                    {"name": "med_signature", "label": "Signature", "type": "signature", "required": True},
                    {"name": "med_date", "label": "Date", "type": "date", "required": True},
                ]},
            ]
        },
    },
    {
        "id": "prior_history",
        "title": "Prior History Form",
        "description": "Mental health history, trauma overview, family of origin, life events",
        "form_type": "system",
        "form_schema": {
            "sections": [
                {"title": "Mental Health History", "fields": [
                    {"name": "diagnoses", "label": "Previous diagnoses", "type": "textarea"},
                    {"name": "hospitalizations", "label": "Any psychiatric hospitalizations?", "type": "textarea"},
                    {"name": "si_history", "label": "History of suicidal ideation or self-harm?", "type": "dropdown",
                     "options": ["No history", "Past ideation (no current)", "Past attempts (no current)", "Prefer not to answer"]},
                ]},
                {"title": "Trauma History", "fields": [
                    {"name": "trauma_categories", "label": "Areas of trauma (check all that apply)", "type": "multiselect",
                     "options": ["Childhood neglect", "Physical abuse", "Emotional abuse", "Sexual abuse",
                                 "Domestic violence", "Loss/grief", "Medical trauma", "Combat/military", "Accident/disaster", "Other"]},
                ]},
                {"title": "Family of Origin", "fields": [
                    {"name": "family_overview", "label": "Brief family of origin overview", "type": "textarea"},
                    {"name": "family_mh", "label": "Family mental health history", "type": "textarea"},
                ]},
                {"title": "Life Events", "fields": [
                    {"name": "significant_events", "label": "Significant life events (losses, moves, career changes, births)", "type": "textarea"},
                    {"name": "legal_history", "label": "Relevant legal history", "type": "textarea"},
                    {"name": "medical_conditions", "label": "Medical conditions affecting wellbeing", "type": "textarea"},
                ]},
                {"title": "Self-Assessment", "fields": [
                    {"name": "attachment_style", "label": "How would you describe your attachment style?", "type": "dropdown",
                     "options": ["Secure", "Anxious", "Avoidant", "Disorganized", "Not sure"]},
                    {"name": "history_signature", "label": "Signature", "type": "signature", "required": True},
                ]},
            ]
        },
    },
    {
        "id": "goals_obstacles",
        "title": "Goals & Current Obstacles Form",
        "description": "Goals at 30/90/365 days, internal and external obstacles, strengths",
        "form_type": "system",
        "form_schema": {
            "sections": [
                {"title": "Goals", "fields": [
                    {"name": "goals_30d", "label": "Top 3 goals for the next 30 days", "type": "textarea", "required": True},
                    {"name": "goals_90d", "label": "Top 3 goals for the next 90 days", "type": "textarea", "required": True},
                    {"name": "goals_1yr", "label": "Top 3 goals for the next year", "type": "textarea"},
                ]},
                {"title": "Obstacles", "fields": [
                    {"name": "internal_obstacles", "label": "Internal obstacles (anxiety, avoidance, beliefs)", "type": "textarea", "required": True},
                    {"name": "external_obstacles", "label": "External obstacles (finances, relationships, work)", "type": "textarea"},
                ]},
                {"title": "Resources & Preferences", "fields": [
                    {"name": "success_criteria", "label": "How will you know you've succeeded?", "type": "textarea"},
                    {"name": "strengths", "label": "Your strengths and resources", "type": "textarea"},
                    {"name": "previous_attempts", "label": "Previous attempts and outcomes", "type": "textarea"},
                    {"name": "coaching_style", "label": "Preferred coaching style", "type": "dropdown",
                     "options": ["Direct and challenging", "Gentle and supportive", "Structured and goal-oriented", "Exploratory and reflective", "No preference"]},
                    {"name": "commitment_level", "label": "Availability and commitment level", "type": "text"},
                    {"name": "support_people", "label": "Support people in your life", "type": "textarea"},
                ]},
            ]
        },
    },
    {
        "id": "group_attention",
        "title": "Group Attention Form",
        "description": "Group member role, individual goals, boundaries, communication preferences",
        "form_type": "system",
        "form_schema": {
            "sections": [
                {"title": "Group Information", "fields": [
                    {"name": "member_name", "label": "Full Name", "type": "text", "required": True},
                    {"name": "group_role", "label": "Role in group", "type": "dropdown",
                     "options": ["Couple partner", "Family member", "Team member", "Other"]},
                    {"name": "individual_goals", "label": "Your individual goals within the group", "type": "textarea", "required": True},
                ]},
                {"title": "Boundaries & Comfort", "fields": [
                    {"name": "boundary_topics", "label": "Topics you'd prefer not to discuss in group", "type": "textarea"},
                    {"name": "dynamics_to_address", "label": "Relationship dynamics you'd like to address", "type": "textarea"},
                    {"name": "conflict_style", "label": "Your conflict style", "type": "dropdown",
                     "options": ["Accommodating", "Avoiding", "Competing", "Compromising", "Collaborating"]},
                ]},
                {"title": "Communication", "fields": [
                    {"name": "comm_preferences", "label": "Communication preferences", "type": "textarea"},
                    {"name": "emergency_contact", "label": "Emergency Contact (name & phone)", "type": "text", "required": True},
                ]},
                {"title": "Consent", "fields": [
                    {"name": "confidentiality_agree", "label": "I agree to maintain group confidentiality", "type": "checkbox", "required": True},
                    {"name": "recording_consent", "label": "I consent to group recording and AI observation", "type": "checkbox"},
                    {"name": "group_signature", "label": "Signature", "type": "signature", "required": True},
                    {"name": "group_date", "label": "Date", "type": "date", "required": True},
                ]},
            ]
        },
    },
    {
        "id": "client_insurance",
        "title": "Client Insurance Form",
        "description": "Insurance provider, policy details, authorization, F-code consent",
        "form_type": "system",
        "form_schema": {
            "sections": [
                {"title": "Primary Insurance", "fields": [
                    {"name": "insurance_provider", "label": "Insurance Provider Name", "type": "text", "required": True},
                    {"name": "policy_number", "label": "Policy Number", "type": "text", "required": True},
                    {"name": "group_number", "label": "Group Number", "type": "text"},
                    {"name": "insurance_phone", "label": "Insurance Phone Number", "type": "text"},
                    {"name": "claims_address", "label": "Claims Address", "type": "textarea"},
                ]},
                {"title": "Subscriber Information", "fields": [
                    {"name": "subscriber_name", "label": "Subscriber Name (if different from client)", "type": "text"},
                    {"name": "subscriber_dob", "label": "Subscriber Date of Birth", "type": "date"},
                    {"name": "subscriber_relationship", "label": "Relationship to Client", "type": "dropdown",
                     "options": ["Self", "Spouse", "Parent", "Other"]},
                ]},
                {"title": "Additional Coverage", "fields": [
                    {"name": "auth_number", "label": "Authorization/Pre-certification Number", "type": "text"},
                    {"name": "eap_info", "label": "EAP Information (if applicable)", "type": "text"},
                    {"name": "secondary_insurance", "label": "Secondary Insurance Details", "type": "textarea"},
                ]},
                {"title": "Provider Information", "fields": [
                    {"name": "provider_npi", "label": "Coach/Provider NPI or Credential Number", "type": "text"},
                ]},
                {"title": "Authorization & Consent", "fields": [
                    {"name": "fcode_consent", "label": "I consent to F-code submission for reimbursement", "type": "checkbox", "required": True},
                    {"name": "info_release", "label": "I authorize release of information to insurer", "type": "checkbox", "required": True},
                    {"name": "insurance_signature", "label": "Signature", "type": "signature", "required": True},
                    {"name": "insurance_date", "label": "Date", "type": "date", "required": True},
                ]},
            ]
        },
    },
]


class FormGenerateRequest(BaseModel):
    template_id: str
    format: str = "pdf"
    client_id: Optional[str] = None
    prefill_data: Optional[Dict] = None


class FormEmailRequest(BaseModel):
    template_id: str
    target_type: str = "client"
    target_id: str = ""
    format: str = "pdf"


class FormCreateRequest(BaseModel):
    description: str
    title: Optional[str] = None


class FormUpdateRequest(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    form_schema: Optional[Dict] = None


@router.get("/templates")
async def list_templates(request: Request, user: Dict = Depends(require_coach)):
    """List all form templates (system + coach custom)."""
    db = getattr(request.app.state, "db_pool", None)

    custom_templates = []
    if db:
        coach_id = user.get("hardware_id", user.get("username", ""))
        async with db.acquire() as conn:
            rows = await conn.fetch(
                """SELECT id, title, description, form_type, created_by_ai, created_at
                   FROM coach_form_templates WHERE coach_id = $1 OR is_system_template = TRUE
                   ORDER BY is_system_template DESC, created_at DESC""",
                coach_id,
            )
            for r in rows:
                custom_templates.append({
                    "id": str(r["id"]),
                    "title": r["title"],
                    "description": r["description"],
                    "form_type": r["form_type"],
                    "created_by_ai": r["created_by_ai"],
                    "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                })

    system = [{"id": t["id"], "title": t["title"], "description": t["description"],
               "form_type": "system", "created_by_ai": False} for t in SYSTEM_TEMPLATES]

    return {"templates": system + custom_templates, "count": len(system) + len(custom_templates)}


@router.get("/templates/{template_id}")
async def get_template(template_id: str, request: Request, user: Dict = Depends(require_coach)):
    """Get a single template with full schema."""
    for t in SYSTEM_TEMPLATES:
        if t["id"] == template_id:
            return t

    db = getattr(request.app.state, "db_pool", None)
    if db:
        async with db.acquire() as conn:
            row = await conn.fetchrow(
                "SELECT id, coach_id, title, description, form_schema, form_type, created_by_ai FROM coach_form_templates WHERE id = $1::uuid",
                template_id,
            )
            if row:
                return {
                    "id": str(row["id"]),
                    "title": row["title"],
                    "description": row["description"],
                    "form_type": row["form_type"],
                    "form_schema": json.loads(row["form_schema"]) if isinstance(row["form_schema"], str) else row["form_schema"],
                    "created_by_ai": row["created_by_ai"],
                }

    raise HTTPException(404, "Template not found")


@router.post("/generate")
async def generate_form(req: FormGenerateRequest, request: Request, user: Dict = Depends(require_coach)):
    """Generate a filled PDF or Excel from a template."""
    template = None
    for t in SYSTEM_TEMPLATES:
        if t["id"] == req.template_id:
            template = t
            break

    if not template:
        db = getattr(request.app.state, "db_pool", None)
        if db:
            async with db.acquire() as conn:
                row = await conn.fetchrow(
                    "SELECT form_schema, title FROM coach_form_templates WHERE id = $1::uuid", req.template_id
                )
                if row:
                    schema = json.loads(row["form_schema"]) if isinstance(row["form_schema"], str) else row["form_schema"]
                    template = {"id": req.template_id, "title": row["title"], "form_schema": schema}

    if not template:
        raise HTTPException(404, "Template not found")

    prefill = req.prefill_data or {}
    if req.client_id:
        db = getattr(request.app.state, "db_pool", None)
        if db:
            async with db.acquire() as conn:
                client_row = await conn.fetchrow(
                    "SELECT profile_data FROM users WHERE username = $1 AND role = 'CLIENT'", req.client_id
                )
                if client_row:
                    profile = client_row["profile_data"]
                    if isinstance(profile, str):
                        profile = json.loads(profile)
                    prefill.update({
                        "legal_name": profile.get("name", ""),
                        "client_name": profile.get("name", ""),
                        "member_name": profile.get("name", ""),
                        "email": profile.get("email", ""),
                        "phone": profile.get("phone", ""),
                        "dob": profile.get("dob", ""),
                        "insurance_provider": profile.get("insurance_provider", ""),
                        "policy_number": profile.get("insurance_policy_number", ""),
                        "group_number": profile.get("insurance_group_number", ""),
                    })

    if req.format == "xlsx":
        content = _generate_excel(template, prefill)
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{template["title"]}.xlsx"'},
        )
    else:
        content = _generate_pdf(template, prefill)
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{template["title"]}.pdf"'},
        )


@router.post("/create")
async def create_custom_form(req: FormCreateRequest, request: Request, user: Dict = Depends(require_coach)):
    """AI-powered form creation from natural language description."""
    db = getattr(request.app.state, "db_pool", None)
    if not db:
        raise HTTPException(503, "Database unavailable")

    coach_id = user.get("hardware_id", user.get("username", ""))
    form_schema = await _ai_generate_form(req.description)

    title = req.title or form_schema.get("title", "Custom Form")

    async with db.acquire() as conn:
        row = await conn.fetchrow(
            """INSERT INTO coach_form_templates (coach_id, title, description, form_schema, form_type, created_by_ai)
               VALUES ($1, $2, $3, $4::jsonb, 'ai_generated', TRUE) RETURNING id""",
            coach_id, title, req.description, json.dumps(form_schema),
        )

    return {"id": str(row["id"]), "title": title, "form_schema": form_schema, "status": "created"}


@router.put("/{form_id}")
async def update_form(form_id: str, req: FormUpdateRequest, request: Request, user: Dict = Depends(require_coach)):
    """Update a custom form template."""
    db = getattr(request.app.state, "db_pool", None)
    if not db:
        raise HTTPException(503, "Database unavailable")

    coach_id = user.get("hardware_id", user.get("username", ""))

    async with db.acquire() as conn:
        existing = await conn.fetchrow(
            "SELECT id FROM coach_form_templates WHERE id = $1::uuid AND coach_id = $2", form_id, coach_id
        )
        if not existing:
            raise HTTPException(404, "Form not found")

        updates = []
        params = [form_id]
        idx = 2
        if req.title:
            updates.append(f"title = ${idx}")
            params.append(req.title)
            idx += 1
        if req.description:
            updates.append(f"description = ${idx}")
            params.append(req.description)
            idx += 1
        if req.form_schema:
            updates.append(f"form_schema = ${idx}::jsonb")
            params.append(json.dumps(req.form_schema))
            idx += 1

        if updates:
            updates.append("updated_at = NOW()")
            await conn.execute(
                f"UPDATE coach_form_templates SET {', '.join(updates)} WHERE id = $1::uuid", *params
            )

    return {"status": "updated"}


@router.delete("/{form_id}")
async def delete_form(form_id: str, request: Request, user: Dict = Depends(require_coach)):
    """Delete a custom form template."""
    db = getattr(request.app.state, "db_pool", None)
    if not db:
        raise HTTPException(503, "Database unavailable")

    coach_id = user.get("hardware_id", user.get("username", ""))

    async with db.acquire() as conn:
        result = await conn.execute(
            "DELETE FROM coach_form_templates WHERE id = $1::uuid AND coach_id = $2 AND is_system_template = FALSE",
            form_id, coach_id,
        )

    if "DELETE 0" in result:
        raise HTTPException(404, "Form not found or is a system template")
    return {"status": "deleted"}


@router.post("/email")
async def email_form(req: FormEmailRequest, request: Request, user: Dict = Depends(require_coach)):
    """Email a form to client(s) by individual, family, group, or company."""
    db = getattr(request.app.state, "db_pool", None)
    if not db:
        raise HTTPException(503, "Database unavailable")

    async with db.acquire() as conn:
        if req.target_type == "client":
            rows = await conn.fetch(
                "SELECT username, profile_data->>'email' as email, profile_data->>'name' as name FROM users WHERE username = $1",
                req.target_id,
            )
        elif req.target_type == "family":
            rows = await conn.fetch(
                "SELECT username, profile_data->>'email' as email, profile_data->>'name' as name FROM users WHERE profile_data->>'family_id' = $1",
                req.target_id,
            )
        elif req.target_type == "group":
            rows = await conn.fetch(
                "SELECT username, profile_data->>'email' as email, profile_data->>'name' as name FROM users WHERE profile_data->>'group_id' = $1",
                req.target_id,
            )
        else:
            rows = await conn.fetch(
                "SELECT username, profile_data->>'email' as email, profile_data->>'name' as name FROM users WHERE profile_data->>'company_id' = $1",
                req.target_id,
            )

    recipients = [{"username": r["username"], "email": r["email"], "name": r["name"]} for r in rows if r["email"]]

    if not recipients:
        raise HTTPException(404, "No recipients with email found")

    notify = getattr(request.app.state, "notification_system", None)
    sent_count = 0
    for recipient in recipients:
        try:
            if notify:
                await _send_form_email(notify, recipient, req.template_id)
                sent_count += 1
        except Exception as e:
            logger.warning("Failed to email form to %s: %s", recipient["email"], e)

    return {"status": "sent", "recipients": sent_count, "total_found": len(recipients)}


@router.get("/health")
async def forms_health():
    """Health check."""
    return {"status": "ok", "service": "coach_forms", "system_templates": len(SYSTEM_TEMPLATES)}


def _generate_pdf(template: Dict, prefill: Dict) -> bytes:
    """Generate PDF from template schema with prefilled data."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.75 * inch, bottomMargin=0.75 * inch)
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle("FormTitle", parent=styles["Title"], fontSize=18, spaceAfter=12, textColor=colors.HexColor("#1a1a1a"))
        section_style = ParagraphStyle("SectionTitle", parent=styles["Heading2"], fontSize=14, spaceBefore=18, spaceAfter=6, textColor=colors.HexColor("#333333"))
        field_label_style = ParagraphStyle("FieldLabel", parent=styles["Normal"], fontSize=10, spaceAfter=2, textColor=colors.HexColor("#555555"))
        field_value_style = ParagraphStyle("FieldValue", parent=styles["Normal"], fontSize=11, spaceAfter=8, leftIndent=12)

        elements = []
        elements.append(Paragraph(template.get("title", "Form"), title_style))
        elements.append(Spacer(1, 12))

        schema = template.get("form_schema", {})
        for section in schema.get("sections", []):
            elements.append(Paragraph(section.get("title", ""), section_style))
            for field in section.get("fields", []):
                label = field.get("label", "")
                name = field.get("name", "")
                ftype = field.get("type", "text")
                value = prefill.get(name, field.get("value", ""))

                elements.append(Paragraph(f"<b>{label}:</b>", field_label_style))
                if ftype == "info":
                    elements.append(Paragraph(str(value), field_value_style))
                elif ftype in ("checkbox", "multiselect"):
                    display = "Yes" if value else "[ ]"
                    elements.append(Paragraph(display, field_value_style))
                elif ftype == "signature":
                    elements.append(Paragraph(str(value) if value else "___________________________", field_value_style))
                else:
                    elements.append(Paragraph(str(value) if value else "_____________________", field_value_style))

        doc.build(elements)
        return buffer.getvalue()

    except ImportError:
        logger.warning("reportlab not installed — returning placeholder PDF")
        return b"%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n3 0 obj<</Type/Page/MediaBox[0 0 612 792]/Parent 2 0 R>>endobj\nxref\n0 4\n0000000000 65535 f \n0000000009 00000 n \n0000000058 00000 n \n0000000115 00000 n \ntrailer<</Size 4/Root 1 0 R>>\nstartxref\n190\n%%EOF"


def _generate_excel(template: Dict, prefill: Dict) -> bytes:
    """Generate Excel from template schema with prefilled data."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = template.get("title", "Form")[:31]

        header_font = Font(name="Calibri", size=14, bold=True, color="1A1A1A")
        section_font = Font(name="Calibri", size=12, bold=True, color="333333")
        section_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        label_font = Font(name="Calibri", size=10, color="555555")
        value_font = Font(name="Calibri", size=11)

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 50

        row = 1
        ws.cell(row=row, column=1, value=template.get("title", "Form")).font = header_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 2

        schema = template.get("form_schema", {})
        for section in schema.get("sections", []):
            ws.cell(row=row, column=1, value=section.get("title", "")).font = section_font
            ws.cell(row=row, column=1).fill = section_fill
            ws.cell(row=row, column=2).fill = section_fill
            row += 1

            for field in section.get("fields", []):
                label = field.get("label", "")
                name = field.get("name", "")
                value = prefill.get(name, field.get("value", ""))

                ws.cell(row=row, column=1, value=label).font = label_font
                ws.cell(row=row, column=2, value=str(value) if value else "").font = value_font
                row += 1
            row += 1

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    except ImportError:
        logger.warning("openpyxl not installed — returning placeholder XLSX")
        return b"PK"


async def _ai_generate_form(description: str) -> Dict:
    """Use Azure OpenAI to generate a form schema from natural language."""
    try:
        import httpx

        if not NATE_CHAT_KEY:
            return _fallback_form_schema(description)

        messages = [
            {"role": "system", "content": """You are a clinical form designer. Generate a JSON form schema from the user's description.
Return ONLY valid JSON with this structure:
{
  "title": "Form Title",
  "sections": [
    {
      "title": "Section Name",
      "fields": [
        {"name": "field_key", "label": "Display Label", "type": "text|textarea|checkbox|dropdown|date|signature|multiselect|info", "required": true|false, "options": ["opt1", "opt2"]}
      ]
    }
  ]
}
Use clinical best practices. Include consent and signature fields."""},
            {"role": "user", "content": description},
        ]

        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(NATE_CHAT_URL, json=nate_chat_payload(messages=messages, max_tokens=2000), headers=nate_chat_headers())
            if resp.status_code == 200:
                data = resp.json()
                content = data["choices"][0]["message"]["content"]
                content = content.strip()
                if content.startswith("```"):
                    content = content.split("\n", 1)[1].rsplit("```", 1)[0]
                return json.loads(content)
    except Exception as e:
        logger.warning("AI form generation failed: %s", e)

    return _fallback_form_schema(description)


def _fallback_form_schema(description: str) -> Dict:
    """Fallback form when AI is unavailable."""
    return {
        "title": "Custom Form",
        "sections": [
            {"title": "General Information", "fields": [
                {"name": "client_name", "label": "Full Name", "type": "text", "required": True},
                {"name": "date", "label": "Date", "type": "date", "required": True},
                {"name": "notes", "label": description[:200], "type": "textarea", "required": True},
            ]},
            {"title": "Consent", "fields": [
                {"name": "signature", "label": "Signature", "type": "signature", "required": True},
            ]},
        ],
    }


async def _send_form_email(notify, recipient: Dict, template_id: str):
    """Send form email via notification system."""
    template = None
    for t in SYSTEM_TEMPLATES:
        if t["id"] == template_id:
            template = t
            break
    if not template:
        return

    subject = f"Form: {template['title']} — Sovereign Sanctuary"
    body = f"""<p>Hello {recipient.get('name', 'Client')},</p>
<p>Your coach has sent you the <strong>{template['title']}</strong> form to complete.</p>
<p>Please log in to the Sovereign Sanctuary app to fill it out.</p>
<p>Thank you,<br>Sovereign Sanctuary</p>"""

    if hasattr(notify, "_send_email"):
        await notify._send_email(recipient["email"], subject, body)
