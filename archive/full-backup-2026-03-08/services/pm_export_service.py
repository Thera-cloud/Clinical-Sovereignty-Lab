"""
PM & Business DOJO — Gantt Chart (PDF) and Excel Export Service
================================================================
Extracts structured project data from Dojo conversation history,
then generates either:
  1. A Gantt chart PDF (using fpdf2)
  2. An Excel workbook (using openpyxl)

Used by: bridge_server.py WebSocket handlers (dojo_generate_gantt, dojo_generate_excel)
"""

import io
import json
import os
import uuid
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
from pathlib import Path

from app.services.nate_ai_config import NATE_CHAT_URL, NATE_CHAT_KEY, nate_chat_headers, nate_chat_payload

logger = logging.getLogger(__name__)

# =============================================================================
# AI EXTRACTION — Ask GPT-4o to parse conversation into structured project data
# =============================================================================

EXTRACTION_PROMPT = """You are a project data extractor. Analyze the conversation below and extract a structured project plan.

Return ONLY valid JSON with this exact structure (no markdown, no explanation):
{
  "project_name": "string",
  "project_summary": "1-2 sentence summary",
  "tasks": [
    {
      "id": 1,
      "name": "Task name",
      "start_day": 1,
      "duration_days": 5,
      "assignee": "Person or Team",
      "status": "not_started",
      "dependencies": [],
      "category": "Phase or Category name",
      "priority": "high|medium|low"
    }
  ],
  "milestones": [
    {
      "name": "Milestone name",
      "day": 10
    }
  ],
  "total_days": 30
}

Rules:
- start_day is 1-indexed (Day 1 = project start)
- duration_days must be >= 1
- dependencies is a list of task id integers
- If no specific dates are mentioned, create reasonable estimates
- Extract 5-20 tasks depending on conversation detail
- Group tasks into logical categories/phases
- status should be "not_started" for all tasks
- If conversation doesn't contain project info, create a sample project plan based on any context clues

CONVERSATION:
"""


async def extract_project_data(messages: List[Dict], azure_api_key: str, azure_endpoint: str) -> Dict:
    """
    Send conversation to Azure OpenAI GPT-4o (REST chat completions) 
    to extract structured project data as JSON.
    """
    import aiohttp

    # Build conversation text from messages
    conversation_text = ""
    for msg in messages:
        role = msg.get("role", "user")
        text = msg.get("text", "") or msg.get("content", "")
        conversation_text += f"[{role.upper()}]: {text}\n\n"

    messages = [
        {"role": "system", "content": EXTRACTION_PROMPT},
        {"role": "user", "content": conversation_text},
    ]
    payload = nate_chat_payload(messages, max_tokens=2000)
    payload["response_format"] = {"type": "json_object"}

    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(NATE_CHAT_URL, json=payload, headers=nate_chat_headers(), timeout=aiohttp.ClientTimeout(total=30)) as resp:
                if resp.status != 200:
                    error_text = await resp.text()
                    logger.error(f"Azure API error {resp.status}: {error_text}")
                    # Return a fallback sample project
                    return _fallback_project_data(conversation_text)
                
                result = await resp.json()
                content = result["choices"][0]["message"]["content"]
                return json.loads(content)
    except Exception as e:
        logger.error(f"AI extraction failed: {e}")
        return _fallback_project_data(conversation_text)


def _fallback_project_data(context: str = "") -> Dict:
    """Fallback project data if AI extraction fails."""
    return {
        "project_name": "Project Plan",
        "project_summary": "Auto-generated project plan from Dojo session",
        "tasks": [
            {"id": 1, "name": "Requirements Gathering", "start_day": 1, "duration_days": 5, "assignee": "Team", "status": "not_started", "dependencies": [], "category": "Planning", "priority": "high"},
            {"id": 2, "name": "Design Phase", "start_day": 6, "duration_days": 7, "assignee": "Team", "status": "not_started", "dependencies": [1], "category": "Planning", "priority": "high"},
            {"id": 3, "name": "Development Sprint 1", "start_day": 13, "duration_days": 10, "assignee": "Dev Team", "status": "not_started", "dependencies": [2], "category": "Development", "priority": "high"},
            {"id": 4, "name": "Testing", "start_day": 23, "duration_days": 5, "assignee": "QA Team", "status": "not_started", "dependencies": [3], "category": "Quality", "priority": "medium"},
            {"id": 5, "name": "Deployment", "start_day": 28, "duration_days": 3, "assignee": "DevOps", "status": "not_started", "dependencies": [4], "category": "Release", "priority": "high"},
        ],
        "milestones": [
            {"name": "Design Complete", "day": 12},
            {"name": "Release", "day": 30}
        ],
        "total_days": 30
    }


# =============================================================================
# GANTT CHART PDF — Render project data as a horizontal bar Gantt chart
# =============================================================================

# Sovereign Sanctuary color palette
COLORS = {
    "bg":       (10, 10, 10),      # #0A0A0A
    "card":     (17, 17, 17),      # #111111
    "elevated": (26, 26, 26),      # #1A1A1A
    "gold":     (201, 169, 98),    # #C9A962
    "gold_dim": (139, 115, 85),    # #8B7355
    "cyan":     (78, 205, 196),    # #4ECDC4
    "purple":   (157, 78, 221),    # #9D4EDD
    "red":      (239, 68, 68),     # #EF4444
    "green":    (0, 255, 136),     # #00FF88
    "white":    (255, 255, 255),
    "grey":     (136, 136, 136),   # #888888
    "border":   (37, 37, 37),      # #252525
}

# Category colors for Gantt bars
CATEGORY_COLORS = [
    (78, 205, 196),    # cyan
    (201, 169, 98),    # gold
    (157, 78, 221),    # purple
    (0, 255, 136),     # green
    (255, 149, 0),     # orange
    (239, 68, 68),     # red
    (100, 149, 237),   # cornflower
    (255, 215, 0),     # bright gold
]


def generate_gantt_pdf(project_data: Dict) -> bytes:
    """
    Generate a Gantt chart PDF from structured project data.
    Returns PDF as bytes.
    """
    from fpdf import FPDF

    tasks = project_data.get("tasks", [])
    milestones = project_data.get("milestones", [])
    project_name = project_data.get("project_name", "Project Plan")
    total_days = project_data.get("total_days", 30)
    if not tasks:
        tasks = _fallback_project_data()["tasks"]
        total_days = 30

    # Ensure total_days covers all tasks
    max_end = max((t.get("start_day", 1) + t.get("duration_days", 1) - 1) for t in tasks) if tasks else 30
    total_days = max(total_days, max_end + 2)

    # Layout constants
    page_w = 297  # A4 landscape width mm
    page_h = 210  # A4 landscape height mm
    margin = 15
    header_h = 30
    label_w = 65  # task name column width
    row_h = 8
    timeline_x = margin + label_w + 5
    timeline_w = page_w - timeline_x - margin
    day_w = timeline_w / total_days

    # Build category color map
    categories = list(dict.fromkeys(t.get("category", "General") for t in tasks))
    cat_colors = {}
    for i, cat in enumerate(categories):
        cat_colors[cat] = CATEGORY_COLORS[i % len(CATEGORY_COLORS)]

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()

    # Dark background
    pdf.set_fill_color(*COLORS["bg"])
    pdf.rect(0, 0, page_w, page_h, "F")

    # Title
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(*COLORS["gold"])
    pdf.set_xy(margin, margin)
    pdf.cell(0, 10, project_name, ln=True)

    # Subtitle
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(*COLORS["grey"])
    summary = project_data.get("project_summary", "")
    pdf.set_xy(margin, margin + 10)
    pdf.cell(0, 5, f"{summary}  |  {total_days} days  |  {len(tasks)} tasks", ln=True)

    # Generated stamp
    pdf.set_font("Helvetica", "", 7)
    pdf.set_text_color(*COLORS["gold_dim"])
    pdf.set_xy(page_w - 80, margin)
    pdf.cell(65, 5, f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}", align="R")
    pdf.set_xy(page_w - 80, margin + 5)
    pdf.cell(65, 5, "Sovereign Sanctuary | The Dojo", align="R")

    # Timeline header
    y_start = margin + header_h
    pdf.set_draw_color(*COLORS["border"])
    pdf.set_line_width(0.3)

    # Column header: "Task"
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_text_color(*COLORS["gold"])
    pdf.set_xy(margin, y_start - 8)
    pdf.cell(label_w, 6, "TASK", align="L")

    # Day numbers across the top
    pdf.set_font("Helvetica", "", 6)
    pdf.set_text_color(*COLORS["grey"])
    step = max(1, total_days // 20)  # Show ~20 labels max
    for d in range(0, total_days, step):
        x = timeline_x + d * day_w
        pdf.set_xy(x, y_start - 8)
        pdf.cell(day_w * step, 6, f"D{d+1}", align="L")

    # Separator line
    pdf.set_draw_color(*COLORS["gold_dim"])
    pdf.line(margin, y_start - 1, page_w - margin, y_start - 1)

    # Draw tasks
    y = y_start
    current_category = ""
    tasks_per_page = int((page_h - y_start - margin) / row_h)

    for idx, task in enumerate(tasks):
        # Check if we need a new page
        if y + row_h > page_h - margin:
            pdf.add_page()
            pdf.set_fill_color(*COLORS["bg"])
            pdf.rect(0, 0, page_w, page_h, "F")
            y = margin + 10
            # Re-draw timeline header on new page
            pdf.set_font("Helvetica", "", 6)
            pdf.set_text_color(*COLORS["grey"])
            for d in range(0, total_days, step):
                x = timeline_x + d * day_w
                pdf.set_xy(x, y - 8)
                pdf.cell(day_w * step, 6, f"D{d+1}", align="L")
            pdf.set_draw_color(*COLORS["gold_dim"])
            pdf.line(margin, y - 1, page_w - margin, y - 1)

        cat = task.get("category", "General")
        bar_color = cat_colors.get(cat, CATEGORY_COLORS[0])

        # Category header
        if cat != current_category:
            current_category = cat
            pdf.set_font("Helvetica", "B", 7)
            pdf.set_text_color(*bar_color)
            pdf.set_xy(margin, y)
            pdf.cell(label_w, row_h, cat.upper(), align="L")
            y += row_h

            if y + row_h > page_h - margin:
                pdf.add_page()
                pdf.set_fill_color(*COLORS["bg"])
                pdf.rect(0, 0, page_w, page_h, "F")
                y = margin + 10

        # Task name
        pdf.set_font("Helvetica", "", 7)
        pdf.set_text_color(*COLORS["white"])
        pdf.set_xy(margin + 4, y)
        name = task.get("name", f"Task {task.get('id', idx+1)}")
        if len(name) > 28:
            name = name[:26] + ".."
        pdf.cell(label_w - 4, row_h, name, align="L")

        # Gantt bar
        start = max(0, task.get("start_day", 1) - 1)
        duration = max(1, task.get("duration_days", 1))
        bar_x = timeline_x + start * day_w
        bar_w = duration * day_w

        # Bar background (dim)
        pdf.set_fill_color(bar_color[0] // 4, bar_color[1] // 4, bar_color[2] // 4)
        pdf.rect(bar_x, y + 1, bar_w, row_h - 2, "F")

        # Bar foreground
        pdf.set_fill_color(*bar_color)
        pdf.rect(bar_x, y + 1.5, bar_w, row_h - 3, "F")

        # Duration label on bar
        if bar_w > 12:
            pdf.set_font("Helvetica", "B", 5.5)
            pdf.set_text_color(*COLORS["bg"])
            pdf.set_xy(bar_x + 1, y + 1.5)
            pdf.cell(bar_w - 2, row_h - 3, f"{duration}d", align="L")

        # Priority indicator
        priority = task.get("priority", "medium")
        if priority == "high":
            pdf.set_fill_color(*COLORS["red"])
            pdf.rect(margin + 1, y + 2.5, 2, 3, "F")

        # Gridline
        pdf.set_draw_color(*COLORS["border"])
        pdf.set_line_width(0.1)
        pdf.line(margin, y + row_h, page_w - margin, y + row_h)

        y += row_h

    # Draw milestones
    if milestones:
        for ms in milestones:
            day = ms.get("day", 1) - 1
            x = timeline_x + day * day_w
            # Diamond marker across all rows
            pdf.set_fill_color(*COLORS["gold"])
            pdf.set_draw_color(*COLORS["gold"])
            pdf.set_line_width(0.5)
            pdf.line(x, y_start, x, min(y, page_h - margin))
            # Label
            pdf.set_font("Helvetica", "B", 6)
            pdf.set_text_color(*COLORS["gold"])
            label_y = min(y + 2, page_h - margin - 5)
            pdf.set_xy(x - 15, label_y)
            ms_name = ms.get("name", "Milestone")
            pdf.cell(30, 4, f"* {ms_name}", align="C")

    # Legend
    legend_y = min(y + 12, page_h - margin - 12)
    if legend_y < page_h - margin:
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_text_color(*COLORS["gold"])
        pdf.set_xy(margin, legend_y)
        pdf.cell(0, 5, "LEGEND:", align="L")

        lx = margin + 20
        for cat, color in cat_colors.items():
            if lx + 40 > page_w - margin:
                break
            pdf.set_fill_color(*color)
            pdf.rect(lx, legend_y + 1, 6, 3, "F")
            pdf.set_font("Helvetica", "", 6)
            pdf.set_text_color(*COLORS["grey"])
            pdf.set_xy(lx + 7, legend_y)
            pdf.cell(30, 5, cat)
            lx += 38

        # High priority indicator
        pdf.set_fill_color(*COLORS["red"])
        pdf.rect(lx, legend_y + 1, 3, 3, "F")
        pdf.set_font("Helvetica", "", 6)
        pdf.set_text_color(*COLORS["grey"])
        pdf.set_xy(lx + 5, legend_y)
        pdf.cell(25, 5, "High Priority")

    return pdf.output()


# =============================================================================
# EXCEL WORKBOOK — Generate structured project plan as .xlsx
# =============================================================================

def generate_excel(project_data: Dict) -> bytes:
    """
    Generate an Excel workbook from structured project data.
    Returns .xlsx bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    tasks = project_data.get("tasks", [])
    milestones = project_data.get("milestones", [])
    project_name = project_data.get("project_name", "Project Plan")
    total_days = project_data.get("total_days", 30)
    summary = project_data.get("project_summary", "")

    if not tasks:
        tasks = _fallback_project_data()["tasks"]

    wb = Workbook()

    # ─── Styles ───
    gold_fill = PatternFill(start_color="C9A962", end_color="C9A962", fill_type="solid")
    dark_fill = PatternFill(start_color="111111", end_color="111111", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="000000", size=11)
    title_font = Font(name="Calibri", bold=True, color="C9A962", size=14)
    body_font = Font(name="Calibri", color="333333", size=10)
    gold_font = Font(name="Calibri", bold=True, color="C9A962", size=10)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(wrap_text=True, vertical="center")

    # =========================================================================
    # Sheet 1: Project Overview
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Project Overview"
    ws1.sheet_properties.tabColor = "C9A962"

    ws1["A1"] = project_name
    ws1["A1"].font = title_font
    ws1.merge_cells("A1:D1")

    ws1["A3"] = "Summary:"
    ws1["A3"].font = Font(bold=True, size=10)
    ws1["B3"] = summary
    ws1["B3"].font = body_font
    ws1.merge_cells("B3:D3")

    ws1["A4"] = "Total Duration:"
    ws1["A4"].font = Font(bold=True, size=10)
    ws1["B4"] = f"{total_days} days"
    ws1["B4"].font = body_font

    ws1["A5"] = "Total Tasks:"
    ws1["A5"].font = Font(bold=True, size=10)
    ws1["B5"] = len(tasks)
    ws1["B5"].font = body_font

    ws1["A6"] = "Milestones:"
    ws1["A6"].font = Font(bold=True, size=10)
    ws1["B6"] = len(milestones)
    ws1["B6"].font = body_font

    ws1["A7"] = "Generated:"
    ws1["A7"].font = Font(bold=True, size=10)
    ws1["B7"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws1["B7"].font = body_font

    ws1["A8"] = "Source:"
    ws1["A8"].font = Font(bold=True, size=10)
    ws1["B8"] = "Sovereign Sanctuary | The Dojo"
    ws1["B8"].font = body_font

    # Milestones table
    if milestones:
        ws1["A10"] = "MILESTONES"
        ws1["A10"].font = gold_font
        ws1["A11"] = "Name"
        ws1["A11"].font = header_font
        ws1["A11"].fill = gold_fill
        ws1["B11"] = "Day"
        ws1["B11"].font = header_font
        ws1["B11"].fill = gold_fill

        for i, ms in enumerate(milestones):
            row = 12 + i
            ws1[f"A{row}"] = ms.get("name", "Milestone")
            ws1[f"A{row}"].font = body_font
            ws1[f"A{row}"].border = thin_border
            ws1[f"B{row}"] = ms.get("day", 0)
            ws1[f"B{row}"].font = body_font
            ws1[f"B{row}"].border = thin_border
            ws1[f"B{row}"].alignment = center_align

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 40
    ws1.column_dimensions["C"].width = 15
    ws1.column_dimensions["D"].width = 15

    # =========================================================================
    # Sheet 2: Task List (main data)
    # =========================================================================
    ws2 = wb.create_sheet("Task List")
    ws2.sheet_properties.tabColor = "4ECDC4"

    headers = ["ID", "Task Name", "Category", "Start Day", "Duration (days)", "End Day", "Assignee", "Priority", "Dependencies", "Status"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = gold_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, task in enumerate(tasks, 2):
        start_day = task.get("start_day", 1)
        duration = task.get("duration_days", 1)
        end_day = start_day + duration - 1
        deps = task.get("dependencies", [])
        dep_str = ", ".join(str(d) for d in deps) if deps else "None"

        values = [
            task.get("id", row_idx - 1),
            task.get("name", f"Task {row_idx - 1}"),
            task.get("category", "General"),
            start_day,
            duration,
            end_day,
            task.get("assignee", "TBD"),
            task.get("priority", "medium").capitalize(),
            dep_str,
            task.get("status", "not_started").replace("_", " ").title(),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = body_font
            cell.border = thin_border
            if col_idx in (1, 4, 5, 6):
                cell.alignment = center_align
            elif col_idx == 2:
                cell.alignment = wrap_align

        # Color-code priority
        priority_cell = ws2.cell(row=row_idx, column=8)
        p = task.get("priority", "medium").lower()
        if p == "high":
            priority_cell.font = Font(name="Calibri", color="EF4444", bold=True, size=10)
        elif p == "low":
            priority_cell.font = Font(name="Calibri", color="888888", size=10)

    # Column widths
    col_widths = [6, 35, 18, 12, 14, 10, 18, 12, 15, 16]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Auto-filter
    ws2.auto_filter.ref = f"A1:J{len(tasks) + 1}"

    # =========================================================================
    # Sheet 3: Timeline View (visual)
    # =========================================================================
    ws3 = wb.create_sheet("Timeline")
    ws3.sheet_properties.tabColor = "9D4EDD"

    # Header row: Task name + day columns
    ws3.cell(row=1, column=1, value="Task").font = header_font
    ws3.cell(row=1, column=1).fill = gold_fill
    ws3.cell(row=1, column=1).border = thin_border
    ws3.column_dimensions["A"].width = 30

    for d in range(1, total_days + 1):
        col = d + 1
        cell = ws3.cell(row=1, column=col, value=f"D{d}")
        cell.font = Font(name="Calibri", bold=True, size=8, color="666666")
        cell.alignment = center_align
        cell.border = thin_border
        ws3.column_dimensions[get_column_letter(col)].width = 4

    # Task rows with filled cells for duration
    cat_fill_map = {}
    cat_hex = ["4ECDC4", "C9A962", "9D4EDD", "00FF88", "FF9500", "EF4444", "6495ED", "FFD700"]

    for row_idx, task in enumerate(tasks, 2):
        name = task.get("name", f"Task {row_idx - 1}")
        ws3.cell(row=row_idx, column=1, value=name).font = body_font
        ws3.cell(row=row_idx, column=1).border = thin_border

        cat = task.get("category", "General")
        if cat not in cat_fill_map:
            hex_color = cat_hex[len(cat_fill_map) % len(cat_hex)]
            cat_fill_map[cat] = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

        start = task.get("start_day", 1)
        duration = task.get("duration_days", 1)

        for d in range(start, start + duration):
            if d <= total_days:
                col = d + 1
                cell = ws3.cell(row=row_idx, column=col)
                cell.fill = cat_fill_map[cat]
                cell.border = thin_border

    # Milestone markers
    for ms in milestones:
        day = ms.get("day", 1)
        if day <= total_days:
            col = day + 1
            ms_row = len(tasks) + 3
            cell = ws3.cell(row=ms_row, column=col, value="*")
            cell.font = Font(name="Calibri", bold=True, color="C9A962", size=12)
            cell.alignment = center_align

    if milestones:
        ms_label_row = len(tasks) + 3
        ws3.cell(row=ms_label_row, column=1, value="MILESTONES").font = gold_font

    # Freeze panes (task names stay visible while scrolling timeline)
    ws3.freeze_panes = "B2"

    # =========================================================================
    # Return bytes
    # =========================================================================
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# =============================================================================
# FILE STORAGE — Save generated files for download
# =============================================================================

def save_export_file(file_bytes: bytes, file_type: str, mode: str) -> Dict:
    """
    Save generated file to disk and return metadata for download.
    file_type: 'gantt_pdf' or 'excel'
    Returns: {"file_id": str, "filename": str, "path": str}
    """
    data_dir = Path(os.environ.get("DATA_DIR", "/app/data"))
    export_dir = data_dir / "dojo_exports"
    export_dir.mkdir(parents=True, exist_ok=True)

    file_id = f"DOJO_{mode.upper()}_{file_type.upper()}_{uuid.uuid4().hex[:8].upper()}"

    if file_type == "gantt_pdf":
        filename = f"{file_id}.pdf"
    else:
        filename = f"{file_id}.xlsx"

    file_path = export_dir / filename
    file_path.write_bytes(file_bytes)

    content_type = "application/pdf" if file_type == "gantt_pdf" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    try:
        from app.services.blob_storage import upload_bytes as blob_upload
        blob_upload(
            rel_path=f"dojo_exports/{filename}",
            content=file_bytes, content_type=content_type,
        )
    except Exception as e:
        logger.debug("R2 backup of export skipped: %s", e)

    logger.info(f"Saved export: {filename} ({len(file_bytes)} bytes)")

    return {
        "file_id": file_id,
        "filename": filename,
        "path": str(file_path),
    }
